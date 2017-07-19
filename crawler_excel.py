#read the ARs homepages from the file 'crawler.xlsx'
# obtain all the links from the website, starting from the homepage
#select the links we need (only the links that make the full website)
#write the links in the file 'crawler.xlsx'

from bs4 import BeautifulSoup      #library to recognize objects in the html code
import requests
import urllib.robotparser          #library to check the robots.txt file (links permitted to crawl)
import sys
import openpyxl                     #library to read/write Excel files
import datetime
import os

#define function to check robots.txt file (using urllib.robotparser)
#given the homepage, check if a page can be crawled or not
def check_bots(homepage,page):
    url_check=page+'/'
    rp=urllib.robotparser.RobotFileParser()
    url_bots=homepage+'/robots.txt'
    rp.set_url(url_bots)
    rp.read()
    return(rp.can_fetch('*',url_check))

#define general function that, given a url/page which has already been asked, takes all the 'a' 'href' links from the html code
#it will be used to loop through all the pages of the website, and find all the links
def find_links():
    #used BeautifulSoup to help us recognize 'a' and 'href' from the html code
    soup=BeautifulSoup(web,'html.parser')
    #loop through all the 'a' 'href' to find the links we need
    for x in soup.find_all('a'):
        link=x.get('href')         #from the 'a' get the 'href' where the link is coded
        if link==None:          #not valid links of NoneType are converted into string 'None', to prevent issues
            link='None'
            continue        #if link is None continue to next step of the for loop
        # if the link is an image/pdf add it to links_image_pdf list and continue to the next step of the loop
        if '.' in link:
            if link.rsplit('.', 1)[1].lower() in ['pdf', 'png', 'jpg', 'gif', 'jpeg']:  # if extension is image/pdf
                links_image_pdf.append(link)
                continue  # go to next step in the for loop
        #part of the URL after ? is not considered, it is deleted
        if '?' in link:
            link=link.split('?',1)[0]
        #if the link contains 'http' and domain, it is a valid link that we need (http://www.kession.com)
        if 'http' in link:
            if domain in link[0:len(url)+5]:
                #check if the link is already stored in links_list, if not we need it
                if link not in links_list:
                    #check robots.txt file, are we allowed to crawl the link?
                    bots_allowed=check_bots(url,link)
                    if bots_allowed is True:
                        links_list.append(link)
                    elif bots_allowed is False:
                        not_allowed_links.append(link)
                if link in links_list:
                    links_repeated.append(link)
            else:
                not_links_list.append(link)
        #if the link starts with '/', add it to url and check if it works (/about)
        elif link.startswith('/'):
            link_full=url+link
            #check if link_full is already present in link_list, if not we need it
            if link_full not in links_list:
                #check robots file, are we allowed to crawl the link?
                bots_allowed=check_bots(url,link_full)
                if bots_allowed is True:
                    try:
                        requests.get(link_full)       #check if the link works
                        links_list.append(link_full)
                    except:
                        not_links_list.append(link)
                elif bots_allowed is False:
                    not_allowed_links.append(link_full)
            elif link_full in links_list:
                links_repeated.append(link_full)
        #if the link does not fall in the above cases, it is not a valid link (#)
        else:
            not_links_list.append(link)


#open the Excel file 'crawler.xlsx' where the list of websites are stored
#given each AR's homepage, the links will be found and written in each AR's tab
wb = openpyxl.load_workbook('crawler.xlsx')
wsARs = wb.get_sheet_by_name('ARs')
AR_links_taken=list()

#loop through the ARs listed in the file 'crawler.xlsx'
#for each AR, run the function find_links() and write them in the AR's tab
for x in range(3, len(wsARs['A']) + 1):
    #only selecet the ARs with 'x' in the column E, they are the ARs we will crawl
    if wsARs.cell(row=x,column=5).value!='x':
        continue                #if 'x' is not in the column E go to the next step of the for loop
    ar = wsARs.cell(row=x, column=1).value
    wsARs.cell(row=x,column=6).value=datetime.datetime.today()    #put date of crawling in column F
    AR_links_taken.append(ar)
    print('\n-------------------------------\nCrawling the website of %s.' %(ar))
    url = wsARs.cell(row=x, column=2).value
    #get the text of the homepage, if it does not work get a message
    try:
        web=requests.get(url).text
    except:
        print('Homepage not working (%s)' %(url))
        continue     #if link does not work continue to next step of the for loop
    domain = wsARs.cell(row=x, column=3).value
    keywords=wsARs.cell(row=x, column=4).value
    keywords=keywords.split(', ')   #transform the keywords from string to list
    # create different lists where the links will be stored
    links_list = list()  # links that we will crawl, they make the full website
    links_list.append(url)
    links_repeated = list()  # links that are repeated in links_list
    not_links_list = list()  # links that we are not allowed to crawl
    not_allowed_links = list()  # links that are not part of the website
    checked_pages = list()  # pages from which we have already taken the links 'href'
    links_image_pdf = list()  # valid links that are image or pdf format
    #use the function find_links to find all the links from the homepage
    find_links()
    print('\nLinks taken from:', url)
    #starting from the list of valid links found in the homepage,
    #the function find_links is used to loop through all the pages and find other links
    #the pages checked are recorded in checked_links, to avoid checking multiple times the same page
    #the goal is to find all the pages that constitute the entire AR's website
    for page in links_list:
        #double check if we are allowed from the robots.txt
        bots_allowed=check_bots(url, page)      #using check_bots function
        #if allowed, get the page's code as a text
        if bots_allowed is True:
            try:
                web=requests.get(page).text
                find_links()                    #using the function to find the links in the page
                print('Links taken from:', page)
            except:
                print('Links not taken from:', page)  #if page's url is not working get a message
    #print the lists of links on the screen, it can be used to check if the script works fine
    print('\nList of valid links: ')
    for x in links_list:
        print(x)
    print('\nList of links in image or pdf format. ')
    for x in links_image_pdf:
        print(x)
    print('\nList of links repeated: ')
    for x in links_repeated:
        print(x)
    print('\nList of links that we are not allowed to crawl (see %s/robots.txt). ' % (url))
    for x in not_allowed_links:
        print(x)
    print('\nList of not valid links: ')
    for x in not_links_list:
        print(x)
    # write list of links into Excel file 'crawler.xlsx', see each AR's tab
    print("--------------------\nWriting the links in the AR's tab (Excel file crawler.xlsx).")
    wsCompany = wb.get_sheet_by_name(ar)      #get AR's tab
    #in the AR's tab, move the old links (columns A, B, C) to the right (columns D, E, F)
    for x in range(2, wsCompany.max_row + 1):
        wsCompany.cell(row=x, column=4).value = wsCompany.cell(row=x, column=1).value
        wsCompany.cell(row=x, column=5).value = wsCompany.cell(row=x, column=2).value
        wsCompany.cell(row=x, column=6).value = wsCompany.cell(row=x, column=3).value
        wsCompany.cell(row=x, column=2).value = ''
        wsCompany.cell(row=x, column=1).value = ''
        wsCompany.cell(row=x, column=3).value = ''
    wsCompany['A2'].value = datetime.datetime.today()
    z = 3
    # write the links into the AR's tab
    # also write the length of the html code (number of character)
    for x in links_list:
        wsCompany.cell(row=z, column=1).value = x
        #get the text from the link, if link does not work show a message
        try:
            web=requests.get(x).text
        except:
            wsCompany.cell(row=z, column=2).value='Link not working'
            continue
        wsCompany.cell(row=z, column=2).value = len(web)     #write length of the html code
        #check the keywords
        keywords_found=list()
        for keyword in keywords:
            if keyword.lower() in web.lower():
                keywords_found.append(keyword)
        wsCompany.cell(row=z,column=3).value=', '.join(keywords_found)
        print('Stored %s in crawler.xlsx (AR tab). Keywords found ----> %s' % (x, ', '.join(keywords_found)))
        z += 1
    print('---------------------------')
    # print the length of each list
    print('\nNumber of valid links: ', len(links_list))
    print('Number of image/pdf links: ', len(links_image_pdf))
    print('Number of valid links repeated: ', len(links_repeated))
    print('Number of links that we are not allowed to crawl: ', len(not_allowed_links))
    print('Number of not valid links: ', len(not_links_list))
    print('\n---------------------------------------------\n---------------------------------------------')

#save changes to 'crawler.xlsx' file
print('-----------------------\nThe script has taken the links of:\n')
for x in AR_links_taken:
    print(x)
print('\n-----------------------\nSaving the changes to crawler.xlsx.')
wb.save('crawler.xlsx')

#ask to open crawler.xlsx or exit
while True:
    ask_open=input('Would you like to open (y) the Excel file (crawler.xlsx) or exit (n)? ')
    if ask_open=='y':
        os.startfile('C:\\Users\\Kession9\\PycharmProjects\\Beautifulsoup\\crawler.xlsx')   #open Excel file
        # ask if user wants to exit
        while True:
            ask_exit = input('Exit (y)? ')
            if ask_exit == 'y':
                sys.exit(0)
            else:
                continue
    elif ask_open=='n':
        sys.exit(0)
    else:               #if the answer is not 'y' or 'n' ask again
        continue




